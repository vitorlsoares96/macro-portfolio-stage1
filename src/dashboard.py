"""
Dashboard visual da Fase 1 — monta uma aba "Dashboard" dentro do mesmo
arquivo Excel gerado por exportar_excel.py, com gráficos, um heatmap e um
"cartão" com a leitura atual. Tudo é montado automaticamente pelo Python
— você não precisa criar nenhum gráfico manualmente no Excel depois.

Ideia central deste arquivo (vale entender antes de ler o resto do
código): um gráfico de linha do Excel não sabe pintar trechos da MESMA
linha de cores diferentes dependendo do valor de outra coluna (ex: "verde
quando regime = Expansão, vermelho quando regime = Contração"). O truque
para conseguir esse efeito é a função `separar_por_categoria`: em vez de
1 coluna de valores, criamos N colunas (uma por categoria), cada uma
preenchida só nas linhas em que aquela categoria está ativa, e vazia
(NaN) nas demais linhas. O Excel recebe então N séries — como elas nunca
têm dado na mesma linha, o resultado visual é uma linha só que muda de
cor junto com a categoria. Usamos essa mesma técnica três vezes neste
arquivo: no gráfico de regime, no gráfico de quadrantes (scatter) e no
gráfico de risco tático.

Sobre a ausência de fórmulas do Excel neste arquivo: diferente de uma
planilha financeira "editável" (onde inputs do usuário devem propagar via
fórmula), este dashboard é a saída de um pipeline Python que já fez todo
o cálculo — os números aqui são resultado de um modelo, não premissas que
alguém vai editar depois. Por isso escrevemos valores prontos (como o
resto da exportação já fazia), em vez de recriar as fórmulas de z-score
dentro do Excel.
"""

import pandas as pd

from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONTE = "Calibri"

REGIMES_ORDEM = ["Expansão", "Recuperação", "Desaceleração", "Contração"]
RISCOS_ORDEM = ["Risk-on", "Neutro", "Risk-off"]

CORES_REGIME = {
    "Expansão": "70AD47",       # verde
    "Recuperação": "4472C4",    # azul
    "Desaceleração": "ED7D31",  # laranja
    "Contração": "C00000",      # vermelho
}

CORES_RISCO = {
    "Risk-on": "70AD47",
    "Neutro": "BFBFBF",
    "Risk-off": "C00000",
}

# O gráfico de risco tático é diário — desde 1990 isso são ~9 mil pontos
# por série. O Excel aguenta, mas o gráfico fica visualmente poluído e
# lento para navegar. Por isso recortamos para os últimos N anos aqui;
# o histórico completo continua disponível (sem recorte) na aba
# Modulo_B_Diario, para quem quiser olhar o período inteiro.
ANOS_GRAFICO_RISCO = 5


def separar_por_categoria(valores: pd.Series, categorias: pd.Series, ordem_categorias: list) -> pd.DataFrame:
    """
    Transforma 1 série de valores + 1 série de categorias em N colunas
    (uma por categoria), cada uma preenchida só onde aquela categoria
    está ativa. Ver explicação completa no docstring do módulo.

    `valores.where(categorias == categoria)` mantém o valor original nas
    linhas em que a condição é verdadeira, e substitui por NaN nas
    demais — é o "irmão" do `.mask()` (que faz o oposto).
    """
    return pd.DataFrame({
        categoria: valores.where(categorias == categoria)
        for categoria in ordem_categorias
    })


def montar_dados_auxiliares(resumo_mensal: pd.DataFrame, matriz_combinada: pd.DataFrame) -> dict:
    """
    Monta os 3 DataFrames auxiliares que os gráficos vão ler diretamente
    (cada um vira uma aba auxiliar no Excel):

    1. Aux_Regime — score_ma3 mensal, separado por regime, + uma coluna
       de referência constante em zero (para desenhar a linha do "nível
       neutro" no gráfico).
    2. Aux_Quadrantes — score_ma3 (eixo X) e momentum separado por regime
       (eixo Y), para o scatter que visualiza a regra de classificação
       nível x momentum.
    3. Aux_Risco — score_filtrado diário do Módulo B, separado por risco,
       recortado para os últimos ANOS_GRAFICO_RISCO anos.
    """
    aux_regime = separar_por_categoria(resumo_mensal["score_ma3"], resumo_mensal["regime"], REGIMES_ORDEM)
    aux_regime.insert(0, "zero_referencia", 0.0)

    aux_quadrantes = pd.DataFrame({"score_ma3": resumo_mensal["score_ma3"]})
    momentum_por_regime = separar_por_categoria(resumo_mensal["momentum"], resumo_mensal["regime"], REGIMES_ORDEM)
    aux_quadrantes = aux_quadrantes.join(momentum_por_regime)

    data_corte = matriz_combinada.index.max() - pd.DateOffset(years=ANOS_GRAFICO_RISCO)
    matriz_recente = matriz_combinada.loc[matriz_combinada.index >= data_corte]
    aux_risco = separar_por_categoria(matriz_recente["score_filtrado"], matriz_recente["risco"], RISCOS_ORDEM)

    return {
        "Aux_Regime": aux_regime,
        "Aux_Quadrantes": aux_quadrantes,
        "Aux_Risco": aux_risco,
    }


def montar_tabela_frequencia(matriz_combinada: pd.DataFrame) -> pd.DataFrame:
    """
    Conta quantos dias históricos caíram em cada uma das 12 combinações
    (regime, risco) da matriz de ação — organizado como tabela regime x
    risco, pronta para virar um heatmap. É a mesma pergunta que surgiu
    quando você notou os totais estranhos na tabela dinâmica: aqui dá
    para ver, inclusive, as combinações com contagem zero (que o Excel
    escondia numa tabela dinâmica baseada só na coluna de texto `acao`).
    """
    tabela = pd.crosstab(matriz_combinada["regime"], matriz_combinada["risco"])
    tabela = tabela.reindex(index=REGIMES_ORDEM, columns=RISCOS_ORDEM, fill_value=0)
    return tabela


# ---------------------------------------------------------------------------
# Construção dos gráficos (openpyxl.chart)
# ---------------------------------------------------------------------------

def _grafico_linha_regime(aux_ws, n_linhas) -> LineChart:
    """
    Gráfico de linha do score de regime (mensal), com uma série por
    regime (técnica de separar_por_categoria) + a linha de referência em
    zero. Como cada série só tem dado nos meses do "seu" regime, o
    resultado visual é uma única linha que muda de cor conforme o regime
    muda ao longo do tempo.
    """
    chart = LineChart()
    chart.title = "Regime do ciclo econômico — score suavizado (score_ma3)"
    chart.style = 2
    chart.y_axis.title = "Score (z-score, média móvel 3 meses)"
    chart.x_axis.title = "Data"
    chart.width, chart.height = 30, 12

    # Colunas na aba auxiliar: A=data, B=zero_referencia, C..F=4 regimes.
    dados = Reference(aux_ws, min_col=2, max_col=6, min_row=1, max_row=n_linhas + 1)
    chart.add_data(dados, titles_from_data=True)
    categorias = Reference(aux_ws, min_col=1, min_row=2, max_row=n_linhas + 1)
    chart.set_categories(categorias)

    cores = ["808080"] + [CORES_REGIME[r] for r in REGIMES_ORDEM]
    for serie, cor in zip(chart.series, cores):
        serie.graphicalProperties.line.solidFill = cor
        serie.graphicalProperties.line.width = 22000  # EMU; ~1.7pt
        serie.marker.symbol = "none"
        serie.smooth = False

    # A primeira série (zero_referencia) fica mais fina e tracejada, para
    # não competir visualmente com as séries de regime.
    linha_zero = chart.series[0].graphicalProperties.line
    linha_zero.width = 9000
    linha_zero.dashStyle = "sysDash"

    return chart


def _grafico_quadrantes(aux_ws, n_linhas) -> ScatterChart:
    """
    Scatter de score_ma3 (nível) x momentum, com uma série por regime —
    visualiza diretamente a regra de classificação 2x2 (nível x momentum)
    da Seção 2.2 do documento: cada regime deveria ocupar um quadrante
    visualmente separado dos outros.
    """
    chart = ScatterChart()
    chart.title = "Classificação por quadrante (nível x momentum)"
    chart.x_axis.title = "Score de nível (score_ma3)"
    chart.y_axis.title = "Momentum"
    chart.style = 13
    chart.width, chart.height = 20, 16

    # A aba auxiliar tem: coluna A = data (índice, escrito por to_excel),
    # coluna B = score_ma3 (eixo X), colunas C..F = momentum separado por
    # regime (eixo Y). Ponto que já me confundiu uma vez testando: o
    # índice do DataFrame sempre "rouba" a coluna A quando escrevemos com
    # `index_label` — por isso score_ma3 começa em B, não em A.
    x_valores = Reference(aux_ws, min_col=2, min_row=2, max_row=n_linhas + 1)

    # Colunas C..F = momentum separado por regime (eixo Y, um por série).
    for i, regime in enumerate(REGIMES_ORDEM):
        col = 3 + i
        y_valores = Reference(aux_ws, min_col=col, min_row=1, max_row=n_linhas + 1)
        serie = Series(y_valores, x_valores, title_from_data=True)
        serie.marker.symbol = "circle"
        serie.marker.size = 5
        serie.marker.graphicalProperties.solidFill = CORES_REGIME[regime]
        serie.marker.graphicalProperties.line.noFill = True
        serie.graphicalProperties.line.noFill = True  # só pontos, sem linha conectando
        chart.series.append(serie)

    return chart


def _grafico_linha_risco(aux_ws, n_linhas) -> LineChart:
    """
    Mesma técnica do gráfico de regime, aplicada ao score tático diário
    do Módulo B (score_filtrado), separado por zona de risco.
    """
    chart = LineChart()
    chart.title = f"Risco tático (Módulo B) — últimos {ANOS_GRAFICO_RISCO} anos"
    chart.style = 2
    chart.y_axis.title = "Score filtrado (-2 a +2)"
    chart.x_axis.title = "Data"
    chart.width, chart.height = 30, 12

    dados = Reference(aux_ws, min_col=2, max_col=4, min_row=1, max_row=n_linhas + 1)
    chart.add_data(dados, titles_from_data=True)
    categorias = Reference(aux_ws, min_col=1, min_row=2, max_row=n_linhas + 1)
    chart.set_categories(categorias)

    for serie, risco in zip(chart.series, RISCOS_ORDEM):
        serie.graphicalProperties.line.solidFill = CORES_RISCO[risco]
        serie.graphicalProperties.line.width = 18000
        serie.marker.symbol = "none"
        serie.smooth = False

    return chart


# ---------------------------------------------------------------------------
# Heatmap (formatação condicional) e cartão de status
# ---------------------------------------------------------------------------

def _escrever_heatmap(ws_dashboard, tabela_frequencia: pd.DataFrame, linha_inicio: int, coluna_inicio: int):
    """
    Escreve a tabela regime x risco (contagem de dias) e aplica uma
    `ColorScaleRule` de 3 pontos (vermelho → amarelo → verde) sobre os
    valores — é o jeito nativo do Excel/openpyxl de colorir células de
    acordo com sua magnitude, sem calcular cor nenhuma em Python. Uma
    célula com contagem 0 (a combinação Recuperação + Risk-off, que você
    encontrou) fica no extremo vermelho da escala, bem visível.
    """
    fonte_cabecalho = Font(name=FONTE, bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill("solid", fgColor="404040")

    # Cabeçalho de colunas (riscos)
    ws_dashboard.cell(row=linha_inicio, column=coluna_inicio, value="Regime \\ Risco")
    ws_dashboard.cell(row=linha_inicio, column=coluna_inicio).font = fonte_cabecalho
    ws_dashboard.cell(row=linha_inicio, column=coluna_inicio).fill = preenchimento_cabecalho
    for j, risco in enumerate(RISCOS_ORDEM):
        celula = ws_dashboard.cell(row=linha_inicio, column=coluna_inicio + 1 + j, value=risco)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center")

    # Linhas de dados (regimes)
    for i, regime in enumerate(REGIMES_ORDEM):
        linha = linha_inicio + 1 + i
        celula_rotulo = ws_dashboard.cell(row=linha, column=coluna_inicio, value=regime)
        celula_rotulo.font = Font(name=FONTE, bold=True)
        for j, risco in enumerate(RISCOS_ORDEM):
            valor = int(tabela_frequencia.loc[regime, risco])
            celula = ws_dashboard.cell(row=linha, column=coluna_inicio + 1 + j, value=valor)
            celula.font = Font(name=FONTE)
            celula.alignment = Alignment(horizontal="center")

    # Intervalo só com os valores numéricos (sem cabeçalhos/rótulos), que
    # é o que a ColorScaleRule precisa colorir.
    primeira_linha_dados = linha_inicio + 1
    ultima_linha_dados = linha_inicio + len(REGIMES_ORDEM)
    primeira_coluna_dados = coluna_inicio + 1
    ultima_coluna_dados = coluna_inicio + len(RISCOS_ORDEM)
    intervalo = (
        f"{get_column_letter(primeira_coluna_dados)}{primeira_linha_dados}:"
        f"{get_column_letter(ultima_coluna_dados)}{ultima_linha_dados}"
    )
    regra = ColorScaleRule(
        start_type="min", start_color="F8696B",   # vermelho (menor contagem)
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # amarelo (meio)
        end_type="max", end_color="63BE7B",       # verde (maior contagem)
    )
    ws_dashboard.conditional_formatting.add(intervalo, regra)

    return ultima_linha_dados  # devolve a última linha usada, para posicionar o que vem depois


def _escrever_cartao_status(ws_dashboard, ultima_linha_matriz: pd.Series, linha_inicio: int, coluna_inicio: int):
    """
    Escreve um "cartão" destacado com a leitura mais recente do
    framework: data, regime, risco e ação recomendada — a "manchete" do
    dashboard, com cor de fundo de acordo com o regime atual.
    """
    regime_atual = ultima_linha_matriz["regime"]
    cor_fundo = CORES_REGIME.get(regime_atual, "808080")

    rotulos_valores = [
        ("Leitura atual em", ultima_linha_matriz.name.strftime("%d/%m/%Y")),
        ("Regime (Módulo A)", regime_atual),
        ("Risco (Módulo B)", ultima_linha_matriz["risco"]),
        ("Ação recomendada", ultima_linha_matriz["acao"]),
    ]

    for i, (rotulo, valor) in enumerate(rotulos_valores):
        linha = linha_inicio + i
        celula_rotulo = ws_dashboard.cell(row=linha, column=coluna_inicio, value=rotulo)
        celula_rotulo.font = Font(name=FONTE, bold=True, color="FFFFFF")
        celula_rotulo.fill = PatternFill("solid", fgColor="404040")

        celula_valor = ws_dashboard.cell(row=linha, column=coluna_inicio + 1, value=valor)
        celula_valor.font = Font(name=FONTE, bold=True, size=12, color="FFFFFF")
        celula_valor.fill = PatternFill("solid", fgColor=cor_fundo)
        celula_valor.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_dashboard.column_dimensions[get_column_letter(coluna_inicio)].width = 22
    ws_dashboard.column_dimensions[get_column_letter(coluna_inicio + 1)].width = 55

    return linha_inicio + len(rotulos_valores)


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def montar_dashboard(writer: pd.ExcelWriter, resumo_mensal: pd.DataFrame, matriz_combinada: pd.DataFrame) -> None:
    """
    Função "principal" deste arquivo: recebe o `writer` do Excel já
    aberto (o mesmo usado por exportar_excel.py para as 3 abas de dados)
    e adiciona a ele: 3 abas auxiliares (dados preparados para os
    gráficos) + 1 aba "Dashboard" (cartão de status, heatmap e os 3
    gráficos).

    Precisa ser chamada DENTRO do mesmo bloco `with pd.ExcelWriter(...)`
    de exportar_para_excel — reaproveita o mesmo `writer`, em vez de abrir
    o arquivo de novo, porque o `writer` só grava tudo no disco quando o
    `with` termina.
    """
    dados_auxiliares = montar_dados_auxiliares(resumo_mensal, matriz_combinada)
    for nome_aba, df in dados_auxiliares.items():
        df.to_excel(writer, sheet_name=nome_aba, index_label="data")

    tabela_frequencia = montar_tabela_frequencia(matriz_combinada)

    wb = writer.book
    ws_dashboard = wb.create_sheet("Dashboard", 0)  # posição 0 = primeira aba

    # Cartão de status atual, no canto superior esquerdo.
    proxima_linha = _escrever_cartao_status(
        ws_dashboard, matriz_combinada.iloc[-1], linha_inicio=2, coluna_inicio=2
    )

    # Heatmap regime x risco, logo abaixo do cartão.
    _escrever_heatmap(ws_dashboard, tabela_frequencia, linha_inicio=proxima_linha + 2, coluna_inicio=2)

    # Os 3 gráficos, ancorados em células mais à direita/abaixo para não
    # sobrepor o cartão e o heatmap.
    aux_ws_regime = wb["Aux_Regime"]
    aux_ws_quadrantes = wb["Aux_Quadrantes"]
    aux_ws_risco = wb["Aux_Risco"]

    grafico_regime = _grafico_linha_regime(aux_ws_regime, n_linhas=len(dados_auxiliares["Aux_Regime"]))
    ws_dashboard.add_chart(grafico_regime, "H2")

    grafico_quadrantes = _grafico_quadrantes(aux_ws_quadrantes, n_linhas=len(dados_auxiliares["Aux_Quadrantes"]))
    ws_dashboard.add_chart(grafico_quadrantes, "B18")

    grafico_risco = _grafico_linha_risco(aux_ws_risco, n_linhas=len(dados_auxiliares["Aux_Risco"]))
    ws_dashboard.add_chart(grafico_risco, "H26")

    # As abas auxiliares existem só para alimentar os gráficos — ocultamos
    # para não poluir a visão de quem for abrir o arquivo, mas elas
    # continuam lá (e continuam editáveis) se você quiser conferir os
    # números por trás de algum gráfico.
    for nome_aba in dados_auxiliares:
        wb[nome_aba].sheet_state = "hidden"

    ws_dashboard.sheet_view.showGridLines = False
